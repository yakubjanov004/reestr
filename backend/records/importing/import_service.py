from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from records.models import RegistryRecord, UploadBatch

from .constants import MAX_IMPORT_ERRORS
from .normalizers import normalize_text
from .privacy import apply_privacy_rules
from .record_builder import build_record_kwargs
from .workbook import detect_source_type, find_header_row, is_empty_row, map_row, parse_period


def import_excel_file(*, file_obj, uploaded_by, expected_source_type=None):
    workbook = load_workbook(file_obj, read_only=False, data_only=True)
    worksheet = workbook.active
    file_obj.seek(0)

    header_row_index, headers = find_header_row(worksheet)
    if not headers:
        raise ValueError("To'g'ri formatdagi Excel faylni yuklang. Ustun sarlavhalari shablonga mos bo'lishi kerak.")

    detected_source_type = detect_source_type(headers, worksheet)
    source_type = resolve_source_type(detected_source_type, expected_source_type)
    period_start, period_end = parse_period(worksheet)
    assigned_region = uploaded_by.branch.region if uploaded_by.branch_id else uploaded_by.region
    assigned_branch = uploaded_by.branch

    with transaction.atomic():
        batch = UploadBatch.objects.create(
            uploaded_by=uploaded_by,
            assigned_region=assigned_region,
            assigned_branch=assigned_branch,
            file=file_obj,
            original_filename=getattr(file_obj, "name", "upload.xlsx"),
            source_type=source_type,
            period_start=period_start,
            period_end=period_end,
        )

        rows_in_file = 0
        imported_count = 0
        duplicate_count = 0
        skipped_count = 0
        import_errors = []
        seen_keys = set()

        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            row = map_row(headers, values)
            if is_empty_row(row):
                continue
            rows_in_file += 1

            storage_row = apply_privacy_rules(row, source_type)
            record_kwargs = build_record_kwargs(storage_row, source_type, batch, uploaded_by, row)
            if not record_kwargs["client_name"]:
                skipped_count += 1
                add_import_error(
                    import_errors,
                    row_number=row_number,
                    reason="Mijoz nomi topilmadi.",
                    row=storage_row,
                )
                continue

            dedupe_key = record_kwargs["dedupe_key"]
            if dedupe_key in seen_keys:
                duplicate_count += 1
                add_import_error(
                    import_errors,
                    row_number=row_number,
                    reason="Fayl ichida takroriy yozuv.",
                    row=storage_row,
                )
                continue
            seen_keys.add(dedupe_key)

            try:
                with transaction.atomic():
                    RegistryRecord.objects.create(**record_kwargs)
                imported_count += 1
            except IntegrityError:
                duplicate_count += 1
                add_import_error(
                    import_errors,
                    row_number=row_number,
                    reason="Avval import qilingan yozuv.",
                    row=storage_row,
                )

        batch.rows_in_file = rows_in_file
        batch.imported_count = imported_count
        batch.duplicate_count = duplicate_count
        batch.skipped_count = skipped_count
        batch.import_errors = import_errors
        batch.save(
            update_fields=[
                "rows_in_file",
                "imported_count",
                "duplicate_count",
                "skipped_count",
                "import_errors",
            ]
        )

        if rows_in_file == 0:
            if batch.file:
                batch.file.storage.delete(batch.file.name)
            raise ValueError("Excel faylida ma'lumot topilmadi. Ma'lumotli reestr faylini yuklang.")

    return batch


def resolve_source_type(detected_source_type, expected_source_type):
    if not expected_source_type:
        return detected_source_type

    allowed = {UploadBatch.SourceType.MOBILE, UploadBatch.SourceType.INTERNET}
    if expected_source_type not in allowed:
        raise ValueError("Reestr turi noto'g'ri tanlangan. Mobil raqam yoki Internet ulanish turini tanlang.")

    if detected_source_type == UploadBatch.SourceType.UNKNOWN:
        expected = source_type_label(expected_source_type)
        raise ValueError(f"To'g'ri formatdagi {expected} Excel faylni yuklang.")

    if detected_source_type != expected_source_type:
        expected = source_type_label(expected_source_type)
        actual = source_type_label(detected_source_type)
        raise ValueError(f"Bu {actual} reestri. {actual} turini tanlab yuklang yoki {expected} formatidagi faylni yuklang.")

    return detected_source_type


def source_type_label(source_type):
    labels = {
        UploadBatch.SourceType.MOBILE: "Mobil raqam",
        UploadBatch.SourceType.INTERNET: "Internet ulanish",
        UploadBatch.SourceType.UNKNOWN: "Noma'lum",
    }
    return labels.get(source_type, "Noma'lum")


def add_import_error(errors, *, row_number, reason, row):
    if len(errors) >= MAX_IMPORT_ERRORS:
        return
    errors.append(
        {
            "row": row_number,
            "reason": reason,
            "client_name": normalize_text(row.get("client_name")),
            "identifier": normalize_text(
                row.get("phone_number") or row.get("internet_login") or row.get("request_number")
            ),
        }
    )
