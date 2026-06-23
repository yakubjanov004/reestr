import hashlib
import hmac
import re
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import RegistryRecord, UploadBatch


HEADER_ALIASES = {
    "standard": ["стандарт"],
    "region": ["регион"],
    "dealer": ["дилер"],
    "trade_point": ["торговая точка"],
    "client_name": ["наименование клиента"],
    "document_type": ["тип документа"],
    "document_number": ["серия и номер документа"],
    "birth_date": ["дата рождения"],
    "tariff_plan": ["тарифный план"],
    "phone_number": ["номер телефона"],
    "sim_card_number": ["номер sim-карты (icc)", "номер sim-карты"],
    "contract_number": ["номер контракта"],
    "connection_date": ["дата подключения"],
    "operator_full_name": ["фио оператора"],
    "operator_login": ["логин оператора", "логин"],
    "payment_amount": ["сумма оплат в день подключения", "сумма оплат"],
    "status": ["статус заявки"],
    "identification_method": ["метод идентификации абонента"],
    "technology": ["технология"],
    "internet_login": ["логин интернет"],
    "ip_phone": ["ip-телефрон", "ip-телефон"],
    "account_number": ["лицевой счёт", "лицевой счет"],
    "modem_model": ["модель модема"],
    "modem_serial": ["s/n модема"],
    "modem_cost": ["стоимость модема"],
    "transfer_type": ["тип передачи"],
    "contract_date": ["дата договора"],
    "error_text": ["текст ошибки"],
    "request_number": ["№ заявки", "номер заявки"],
}

DATETIME_FIELDS = {"connection_date", "contract_date"}
DATE_FIELDS = {"birth_date"}
DECIMAL_FIELDS = {"payment_amount", "modem_cost"}
MASKED_FIELDS = {
    UploadBatch.SourceType.MOBILE: {"client_name", "phone_number", "operator_full_name"},
    UploadBatch.SourceType.INTERNET: {"client_name", "internet_login", "operator_full_name"},
}
PART_MASK_FIELDS = {"client_name"}
SUFFIX_DIGIT_MASK_FIELDS = {"phone_number", "internet_login"}
RED_FIELDS = {
    UploadBatch.SourceType.MOBILE: {
        "standard",
        "trade_point",
        "document_type",
        "document_number",
        "birth_date",
        "sim_card_number",
        "contract_number",
        "operator_login",
    },
    UploadBatch.SourceType.INTERNET: {
        "trade_point",
        "document_type",
        "document_number",
        "birth_date",
        "technology",
        "ip_phone",
        "account_number",
        "modem_serial",
        "operator_login",
        "error_text",
    },
}
MAX_IMPORT_ERRORS = 100


def import_excel_file(*, file_obj, uploaded_by, expected_source_type=None):
    workbook = load_workbook(file_obj, read_only=False, data_only=True)
    worksheet = workbook.active
    file_obj.seek(0)

    header_row_index, headers = find_header_row(worksheet)
    if not headers:
        raise ValueError("Excel faylida kerakli sarlavha qatori topilmadi.")

    detected_source_type = detect_source_type(headers, worksheet)
    source_type = resolve_source_type(detected_source_type, expected_source_type)
    period_start, period_end = parse_period(worksheet)

    with transaction.atomic():
        batch = UploadBatch.objects.create(
            uploaded_by=uploaded_by,
            file=file_obj,
            original_filename=getattr(file_obj, "name", "upload.xlsx"),
            source_type=source_type,
            period_start=period_start,
            period_end=period_end,
        )

        rows_in_file = 0
        imported_count = 0
        duplicate_count = 0
        skipped_count = 0
        import_errors = []
        seen_keys = set()

        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            row = map_row(headers, values)
            if is_empty_row(row):
                continue
            rows_in_file += 1

            storage_row = apply_privacy_rules(row, source_type)
            record_kwargs = build_record_kwargs(storage_row, source_type, batch, uploaded_by, row)
            if not record_kwargs["client_name"]:
                skipped_count += 1
                add_import_error(
                    import_errors,
                    row_number=row_number,
                    reason="Mijoz nomi topilmadi.",
                    row=storage_row,
                )
                continue

            dedupe_key = record_kwargs["dedupe_key"]
            if dedupe_key in seen_keys:
                duplicate_count += 1
                add_import_error(
                    import_errors,
                    row_number=row_number,
                    reason="Fayl ichida takroriy yozuv.",
                    row=storage_row,
                )
                continue
            seen_keys.add(dedupe_key)

            try:
                with transaction.atomic():
                    RegistryRecord.objects.create(**record_kwargs)
                imported_count += 1
            except IntegrityError:
                duplicate_count += 1
                add_import_error(
                    import_errors,
                    row_number=row_number,
                    reason="Avval import qilingan yozuv.",
                    row=storage_row,
                )

        batch.rows_in_file = rows_in_file
        batch.imported_count = imported_count
        batch.duplicate_count = duplicate_count
        batch.skipped_count = skipped_count
        batch.import_errors = import_errors
        batch.save(
            update_fields=[
                "rows_in_file",
                "imported_count",
                "duplicate_count",
                "skipped_count",
                "import_errors",
            ]
        )

        if rows_in_file == 0:
            if batch.file:
                batch.file.storage.delete(batch.file.name)
            raise ValueError("Excel faylida ma'lumot topilmadi.")

    return batch


def resolve_source_type(detected_source_type, expected_source_type):
    if not expected_source_type:
        return detected_source_type

    allowed = {UploadBatch.SourceType.MOBILE, UploadBatch.SourceType.INTERNET}
    if expected_source_type not in allowed:
        raise ValueError("Excel turi noto'g'ri tanlangan.")

    if detected_source_type == UploadBatch.SourceType.UNKNOWN:
        return expected_source_type

    if detected_source_type != expected_source_type:
        expected = source_type_label(expected_source_type)
        actual = source_type_label(detected_source_type)
        raise ValueError(f"Tanlangan tur {expected}, lekin fayl {actual} formatida.")

    return detected_source_type


def source_type_label(source_type):
    labels = {
        UploadBatch.SourceType.MOBILE: "Mobil raqam",
        UploadBatch.SourceType.INTERNET: "Internet ulanish",
        UploadBatch.SourceType.UNKNOWN: "Noma'lum",
    }
    return labels.get(source_type, "Noma'lum")


def add_import_error(errors, *, row_number, reason, row):
    if len(errors) >= MAX_IMPORT_ERRORS:
        return
    errors.append(
        {
            "row": row_number,
            "reason": reason,
            "client_name": normalize_text(row.get("client_name")),
            "identifier": normalize_text(
                row.get("phone_number") or row.get("internet_login") or row.get("request_number")
            ),
        }
    )


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def find_header_row(worksheet):
    for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        normalized = [normalize_header(value) for value in row]
        if "наименование клиента" in normalized:
            headers = {}
            for col_index, header in enumerate(normalized):
                field_name = field_for_header(header)
                if field_name:
                    headers[col_index] = field_name
            return index, headers
    return None, {}


def field_for_header(header):
    if not header:
        return None
    for field_name, aliases in HEADER_ALIASES.items():
        if header in aliases:
            return field_name
    return None


def detect_source_type(headers, worksheet):
    fields = set(headers.values())
    if {"phone_number", "sim_card_number"} & fields:
        return UploadBatch.SourceType.MOBILE
    if {"internet_login", "account_number", "technology"} & fields:
        return UploadBatch.SourceType.INTERNET

    title = normalize_text(worksheet.cell(row=1, column=1).value).lower()
    if "шпд" in title:
        return UploadBatch.SourceType.INTERNET
    if "gsm" in title:
        return UploadBatch.SourceType.MOBILE
    return UploadBatch.SourceType.UNKNOWN


def parse_period(worksheet):
    text = normalize_text(worksheet.cell(row=2, column=1).value)
    matches = re.findall(r"\d{2}\.\d{2}\.\d{4}(?:\s+\d{2}:\d{2}:\d{2})?", text)
    if len(matches) < 2:
        return None, None
    return parse_datetime(matches[0]), parse_datetime(matches[1])


def map_row(headers, values):
    row = {}
    for col_index, field_name in headers.items():
        if col_index < len(values):
            row[field_name] = values[col_index]
    return row


def apply_privacy_rules(row, source_type):
    red_fields = RED_FIELDS.get(source_type, set())
    masked_fields = MASKED_FIELDS.get(source_type, set())
    cleaned = {}
    for field_name, value in row.items():
        if field_name in red_fields:
            continue
        if field_name in masked_fields:
            cleaned[field_name] = mask_field_value(field_name, value)
        else:
            cleaned[field_name] = value
    return cleaned


def mask_field_value(field_name, value):
    if field_name in PART_MASK_FIELDS:
        return mask_parts(value)
    if field_name in SUFFIX_DIGIT_MASK_FIELDS:
        return mask_digit_suffix(value)
    return mask_value(value)


def mask_parts(value):
    text = normalize_text(value)
    if not text:
        return ""
    return " ".join(mask_value(part) for part in text.split())


def mask_value(value):
    text = normalize_text(value)
    if not text:
        return ""
    return f"{text[:3]}***"


def mask_digit_suffix(value):
    text = normalize_text(value)
    if not text:
        return ""

    digit_positions = [match.start() for match in re.finditer(r"\d", text)]
    if len(digit_positions) < 3:
        return mask_value(text)

    suffix_start = digit_positions[-3]
    suffix_digits = "".join(text[position] for position in digit_positions[-3:])
    return f"{'*' * suffix_start}{suffix_digits}"


def remask_existing_value(field_name, value):
    text = normalize_text(value)
    if not text:
        return ""
    if field_name in PART_MASK_FIELDS:
        return " ".join(remask_existing_part(part) for part in text.split())
    if field_name in SUFFIX_DIGIT_MASK_FIELDS:
        if text.startswith("***"):
            return text
        if text.endswith("***"):
            return text
        return mask_digit_suffix(text)
    return remask_existing_part(text)


def remask_existing_part(text):
    if text.endswith("***"):
        return text
    return mask_value(text)


def sanitize_record_instance(record):
    red_fields = RED_FIELDS.get(record.source_type, set())
    masked_fields = MASKED_FIELDS.get(record.source_type, set())

    for field_name in red_fields:
        if hasattr(record, field_name):
            field = record._meta.get_field(field_name)
            setattr(record, field_name, None if field.null else "")

    for field_name in masked_fields:
        if hasattr(record, field_name):
            current = getattr(record, field_name)
            if current:
                setattr(record, field_name, remask_existing_value(field_name, current))

    raw_data = {}
    for field_name, value in (record.raw_data or {}).items():
        if field_name in red_fields:
            continue
        raw_data[field_name] = remask_existing_value(field_name, value) if field_name in masked_fields else value
    record.raw_data = raw_data
    return record


def is_empty_row(row):
    return not any(normalize_text(value) for value in row.values())


def build_record_kwargs(row, source_type, batch, uploaded_by, dedupe_row=None):
    raw_data = {key: serialize_cell(value) for key, value in row.items()}
    kwargs = {
        "source_type": source_type,
        "upload_batch": batch,
        "uploaded_by": uploaded_by,
        "raw_data": raw_data,
    }

    for field_name in HEADER_ALIASES:
        if field_name in DATETIME_FIELDS:
            kwargs[field_name] = parse_datetime(row.get(field_name))
        elif field_name in DATE_FIELDS:
            kwargs[field_name] = parse_date(row.get(field_name))
        elif field_name in DECIMAL_FIELDS:
            kwargs[field_name] = parse_decimal(row.get(field_name))
        else:
            kwargs[field_name] = normalize_text(row.get(field_name))

    kwargs["dedupe_key"] = build_dedupe_key(source_type, dedupe_row or kwargs)
    return kwargs


def serialize_cell(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return normalize_text(value)


def parse_date(value):
    parsed = parse_datetime(value)
    if parsed:
        return parsed.date()
    return None


def parse_datetime(value):
    if isinstance(value, datetime):
        return ensure_aware(value)
    if isinstance(value, date):
        return ensure_aware(datetime.combine(value, time.min))
    text = normalize_text(value)
    if not text:
        return None

    formats = (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    )
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            return ensure_aware(parsed)
        except ValueError:
            continue
    return None


def ensure_aware(value):
    if timezone.is_aware(value):
        return value
    return timezone.make_aware(value, timezone.get_current_timezone())


def parse_decimal(value):
    text = normalize_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def build_dedupe_key(source_type, data):
    identity_parts = []
    if data.get("contract_number"):
        identity_parts = ["contract", data["contract_number"]]
    elif data.get("request_number"):
        identity_parts = ["request", data["request_number"]]
    elif source_type == UploadBatch.SourceType.MOBILE and (
        data.get("phone_number") or data.get("sim_card_number")
    ):
        identity_parts = ["mobile", data.get("phone_number"), data.get("sim_card_number")]
    elif source_type == UploadBatch.SourceType.INTERNET and (
        data.get("internet_login") or data.get("account_number")
    ):
        identity_parts = ["internet", data.get("internet_login"), data.get("account_number")]
    else:
        identity_parts = [
            "fallback",
            source_type,
            data.get("document_number"),
            data.get("client_name"),
            serialize_cell(data.get("connection_date") or data.get("contract_date")),
        ]

    normalized = "|".join(normalize_text(part).lower() for part in identity_parts if part)
    secret = settings.SECRET_KEY.encode("utf-8")
    return hmac.new(secret, normalized.encode("utf-8"), hashlib.sha256).hexdigest()
