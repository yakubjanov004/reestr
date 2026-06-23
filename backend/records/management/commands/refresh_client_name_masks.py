from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from records.models import RegistryRecord, UploadBatch
from records.services import (
    MASKED_FIELDS,
    build_dedupe_key,
    find_header_row,
    is_empty_row,
    map_row,
    mask_field_value,
)


class Command(BaseCommand):
    help = "Refresh masked fields from original uploaded Excel files."

    def handle(self, *args, **options):
        updated = 0
        skipped_files = 0

        for batch in UploadBatch.objects.filter(imported_count__gt=0).order_by("id"):
            if not batch.file or not Path(batch.file.path).exists():
                skipped_files += 1
                continue

            workbook = load_workbook(batch.file.path, read_only=False, data_only=True)
            worksheet = workbook.active
            header_row_index, headers = find_header_row(worksheet)
            if not headers:
                skipped_files += 1
                continue

            records_by_key = {
                record.dedupe_key: record
                for record in RegistryRecord.objects.filter(upload_batch=batch)
            }
            touched_keys = set()

            for values in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
                row = map_row(headers, values)
                if is_empty_row(row):
                    continue

                dedupe_key = build_dedupe_key(batch.source_type, row)
                if dedupe_key in touched_keys:
                    continue
                touched_keys.add(dedupe_key)

                record = records_by_key.get(dedupe_key)
                if not record:
                    continue

                masked_fields = MASKED_FIELDS.get(batch.source_type, set())
                changed_fields = []
                original_raw_data = record.raw_data or {}
                raw_data = dict(original_raw_data)

                for field_name in masked_fields:
                    if field_name not in row:
                        continue
                    masked_value = mask_field_value(field_name, row.get(field_name))
                    if not masked_value:
                        continue
                    if getattr(record, field_name, "") != masked_value:
                        setattr(record, field_name, masked_value)
                        changed_fields.append(field_name)
                    if raw_data.get(field_name) != masked_value:
                        raw_data[field_name] = masked_value

                if not changed_fields and raw_data == original_raw_data:
                    continue

                record.raw_data = raw_data
                update_fields = sorted(set(changed_fields + ["raw_data"]))
                record.save(update_fields=update_fields)
                updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Masked fields refreshed for {updated} records. "
                f"Skipped files: {skipped_files}."
            )
        )
